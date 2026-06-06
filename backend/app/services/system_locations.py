"""System-wide locations: CRUD, reorder, and XLSX export/import.

Tenant-agnostic — used by super_admin to manage the canonical location list
that backs the template editor's entity picker.
"""
from __future__ import annotations

from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from ..models import SystemLocation
from ..schemas.system_location import (
    SystemLocationImportSummary,
    SystemLocationIn,
)


HEADER_LABELS = ["מזהה", "שם", "קוד", "פעיל"]


# ---------- CRUD ----------

def create(db: Session, body: SystemLocationIn) -> SystemLocation:
    name = body.name.strip()
    if not name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Name is required"
        )
    if db.query(SystemLocation).filter(SystemLocation.name == name).first():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Location '{name}' already exists",
        )
    next_order = (
        db.query(SystemLocation).count()
    )
    loc = SystemLocation(
        name=name,
        code=(body.code or "").strip() or None,
        is_active=body.is_active,
        sort_order=next_order,
    )
    db.add(loc)
    db.commit()
    db.refresh(loc)
    return loc


def update(db: Session, loc: SystemLocation, body: SystemLocationIn) -> SystemLocation:
    name = body.name.strip()
    if not name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Name is required"
        )
    if name != loc.name:
        existing = (
            db.query(SystemLocation)
            .filter(SystemLocation.name == name, SystemLocation.id != loc.id)
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Location '{name}' already exists",
            )
    loc.name = name
    loc.code = (body.code or "").strip() or None
    loc.is_active = body.is_active
    db.commit()
    db.refresh(loc)
    return loc


def reorder(db: Session, ids: list[int]) -> None:
    """Renumber every active row's sort_order to its position in `ids` (0-based)."""
    rows = db.query(SystemLocation).all()
    by_id = {r.id: r for r in rows}
    for loc_id in ids:
        if loc_id not in by_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"SystemLocation {loc_id} not found",
            )
    for idx, loc_id in enumerate(ids):
        by_id[loc_id].sort_order = idx
    db.commit()


# ---------- XLSX ----------

def export_xlsx(db: Session) -> BytesIO:
    """Build an XLSX in-memory containing every row in sort_order."""
    wb = Workbook()
    ws = wb.active
    ws.title = "מיקומי מערכת"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    for col, label in enumerate(HEADER_LABELS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    rows = (
        db.query(SystemLocation)
        .order_by(SystemLocation.sort_order, SystemLocation.id)
        .all()
    )
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.id)
        ws.cell(row=i, column=2, value=r.name)
        ws.cell(row=i, column=3, value=r.code or "")
        ws.cell(row=i, column=4, value="כן" if r.is_active else "לא")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_bool(v: object) -> bool:
    """Truthy values from a spreadsheet cell. Blank → True (default to active)."""
    if v is None:
        return True
    s = str(v).strip().lower()
    if not s:
        return True
    return s not in ("0", "false", "no", "לא", "כבוי", "off")


def import_xlsx(db: Session, fileobj: BytesIO) -> SystemLocationImportSummary:
    """Replace the entire system_locations table with the contents of an XLSX.

    Full-replacement semantics: every existing row is deleted, then the file's
    rows are inserted fresh with sort_order matching their position in the file.
    The id column in the file is ignored — new auto-increment ids are assigned.
    This avoids the UNIQUE-constraint headaches you get with row-by-row UPDATEs
    when names have been shuffled between rows.

    Columns expected (sheet 1, row 1 is headers and skipped):
      A=id (ignored), B=name, C=code, D=is_active
    Blank rows are skipped. Duplicate names within the file are reported and
    only the first occurrence is kept.
    """
    summary = SystemLocationImportSummary()
    try:
        wb = load_workbook(filename=fileobj, data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Couldn't read XLSX file: {e}",
        )
    ws = wb.active

    # Stage 1: parse the file fully into memory. Validate names + dedup.
    parsed: list[tuple[str, str | None, bool]] = []
    seen_names: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = list(row) + [None] * max(0, 4 - len(row))
        _raw_id, raw_name, raw_code, raw_active = cells[:4]

        name = (str(raw_name).strip() if raw_name is not None else "")
        if not name:
            summary.errors.append(f"שורה {row_idx}: חסר שם — דולג")
            continue
        if name in seen_names:
            summary.errors.append(
                f"שורה {row_idx}: שם כפול '{name}' (כבר בשורה {seen_names[name]}) — דולג"
            )
            continue
        seen_names[name] = row_idx

        code = (str(raw_code).strip() if raw_code is not None else "")
        parsed.append((name, code or None, _parse_bool(raw_active)))

    # Stage 2: wipe + rebuild atomically. SQLAlchemy will commit at the end;
    # if any insert fails (shouldn't, since names are pre-deduped), the
    # rollback restores the original state.
    old_count = db.query(SystemLocation).count()
    db.query(SystemLocation).delete()
    db.flush()
    for i, (name, code, is_active) in enumerate(parsed):
        db.add(
            SystemLocation(
                name=name, code=code, is_active=is_active, sort_order=i
            )
        )

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Import failed: {e}",
        )

    summary.created = len(parsed)
    summary.deleted = old_count
    summary.updated = 0
    return summary
